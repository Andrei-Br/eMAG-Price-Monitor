from openpyxl import Workbook
from openpyxl.styles import Border, Alignment, Side
from openpyxl.utils import get_column_letter

#  variables for formatting the cells
thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="ff0000")

#  function for creating the table
def TableHeader():
    wb = Workbook()
    sheet = wb.active

    #  title of the product
    title_cell = sheet['A1']
    title_cell.value = 'Product'
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    #  the store which sells the product
    website_cell = sheet['B1']
    website_cell.value = 'Store'
    website_cell.alignment = Alignment(horizontal="center", vertical="center")
    website_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    #  date and time column
    date_cell = sheet['C1']
    date_cell.value = 'Date'
    date_cell.alignment = Alignment(horizontal="center", vertical="center")
    date_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    #  the price of the product (lowest price)
    price_cell = sheet['D1']
    price_cell.value = 'Sale price'
    price_cell.alignment = Alignment(horizontal="center", vertical="center")
    price_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    #  link to the product
    link_cell = sheet['E1']
    link_cell.value = 'Link'
    link_cell.alignment = Alignment(horizontal="center", vertical="center")
    link_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


    for row_cell in sheet.iter_rows():
        for cell in enumerate(row_cell):
            sheet.column_dimensions[get_column_letter(cell[0]+1)].width\
                                    = len(cell[1].value) + 2

    #  create an Excel file
    wb.save(filename="eMAG Prices.xlsx")

def main():
    TableHeader()

if __name__ == '__main__':
    main()