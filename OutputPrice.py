import datetime
import openpyxl
from CheckPrice import CheckPrice
from openpyxl.styles import Border, Alignment, Side
from openpyxl.utils import get_column_letter

thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="ff0000")

links = {'ECAM22.100B': 'https://www.emag.ro/espressor-automat-de-longhi'
                       '-magnifica-s-ecam-22-110b-1450w-15-bar-1-8-l-negru-'
                       'ecam22110b/pd/DG0CLBBBM/',
         'ECAM21.117Wh': 'https://www.emag.ro/espressor-automat-de-longhi'
                        '-1450w-15-bar-rasnita-cafea-integrata-alb-ecam-'
                        '21-117-wh/pd/D7JTB0BBM/?ref=',
         'ECAM22.110SB': 'https://www.emag.ro/espressor-automat-de-longhi'
                         '-ecam-22-110-sb-145-0w-15-bar-1-8-l-negru-argintiu-'
                         'ecam-22-110-sb/pd/D9P7PCBBM/?ref=',
         'EP3510': 'https://www.emag.ro/espressor-automat-philips-15-bari-1-8-'
                  'l-sistem-aquaclean-sistem-spumare-a-laptelui-5-setari-'
                  'intensitate-optiune-cafea-macinata-negru-ep3510-00/pd/'
                  'DJ3KYDBBM/?ref=hdr-favorite_products',
         'EP3221': 'https://www.emag.ro/espressor-automat-philips-sistem-de-'
                  'spumare-a-laptelui-4-bauturi-filtru-aquaclean-rasnita-'
                  'ceramica-optiune-cafea-macinata-ecran-tactil-negru-ep3221-'
                  '40/pd/DMYF31BBM/'
         }

def writeExcel():

    # your own path to the Excel file
    path = "D:\Programare\github\emag_price_monitor\eMAG Prices.xlsx"
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    row = sheet.max_row + 1

    #  for each product
    for key in links:
        col = 1

        #  for each product in the url list
        product = CheckPrice((links[key]))

        #  print the name
        title_cell = sheet[get_column_letter(col) + str(row)]
        title_cell.value = product['title']
        col += 1

        #  print the store
        store_cell = sheet[get_column_letter(col) + str(row)]
        store_cell.value = product['store']
        if len(store_cell.value) == 0:
            store_cell.value = 'Partner'
        col += 1

        #  print the time
        date_cell = sheet[get_column_letter(col) + str(row)]
        todays_date = str(datetime.datetime.now().strftime("%d-%m-%Y ; %H:%M"))
        date_cell.value = todays_date
        col += 1

        #  print the price
        price_cell = sheet[get_column_letter(col) + str(row)]
        price_cell.value = str(product['newPrice']) + ' RON'
        col += 1

        #  print the link
        link_cell = sheet[get_column_letter(col) + str(row)]
        link_cell.value = links[key]

        #  Go on the next row after finishing a product
        row += 1

    #  formatting the cells
    for row_cell in sheet.iter_rows():
        for cell in enumerate(row_cell):
            #  ADJUSTING THE WIDTH OF EACH CELL (LENGTH OF THE TEXT + 5
            #  CHARACTERS
            sheet.column_dimensions[get_column_letter(cell[0] + 1)].width \
                = len(cell[1].value) + 10
            """centering each cell"""
            cell[1].alignment = Alignment(horizontal="center",
                                          vertical="center")
            """bordering each cell"""
            cell[1].border = Border(top=thin, left=thin, right=thin,
                                    bottom=thin)

    #  bordering after each sequence of prices
    product_number = len(links)
    for row in range(1, sheet.max_row, product_number):
        for col in range(1, sheet.max_column + 1):
            last_row = sheet[get_column_letter(col) + str(row)]
            last_row.border = Border(top=thin, left=thin, right=thin,
                                     bottom=double)

    wb.save(filename="eMAG Prices.xlsx")

if __name__ == '__main__':
    writeExcel()
